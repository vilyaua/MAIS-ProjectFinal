"""Document parsers for Excel invoices and PDF packing lists."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from .exceptions import ParsingError, UnsupportedFileTypeError
from .models import Item
from .numeric import parse_decimal

SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
SUPPORTED_PDF_EXTENSIONS = {".pdf"}

_CODE_HEADERS = {
    "itemcode",
    "itemno",
    "itemnumber",
    "artikul",
    "article",
    "articleno",
    "sku",
    "code",
    "model",
    "partno",
    "productcode",
    "货号",
    "款号",
    "产品编号",
    "商品编码",
    "编码",
    "型号",
    "品号",
    "артикул",
    "артикулы",
}
_QUANTITY_HEADERS = {
    "quantity",
    "qty",
    "q'ty",
    "pcs",
    "ctns",
    "amount",
    "数量",
    "件数",
    "个数",
    "箱数",
    "количество",
}
_PRICE_HEADERS = {
    "price",
    "unitprice",
    "unitcost",
    "unitvalue",
    "单价",
    "价格",
    "单价usd",
    "usd",
    "цена",
}
_DESCRIPTION_HEADERS = {"description", "name", "product", "品名", "描述", "产品名称", "наименование"}


class DocumentParser:
    """Parser dispatching to Excel or PDF parsing based on file extension."""

    def parse(self, file_path: str | Path) -> list[Item]:
        """Parse document items from a supported file path."""
        path = Path(file_path)
        if not path.exists():
            raise ParsingError(f"Input file does not exist: {path}.")
        if not path.is_file():
            raise ParsingError(f"Input path is not a file: {path}.")

        suffix = path.suffix.lower()
        if suffix in SUPPORTED_EXCEL_EXTENSIONS:
            return self.parse_excel(path)
        if suffix in SUPPORTED_PDF_EXTENSIONS:
            return self.parse_pdf(path)
        raise UnsupportedFileTypeError(
            f"Unsupported file type {suffix or '<none>'!r}. Supported types: "
            f"{', '.join(sorted(SUPPORTED_EXCEL_EXTENSIONS | SUPPORTED_PDF_EXTENSIONS))}."
        )

    def parse_excel(self, file_path: str | Path) -> list[Item]:
        """Parse invoice or packing-list rows from an Excel workbook."""
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency normally installed
            raise ParsingError("Excel parsing requires the 'openpyxl' package.") from exc

        try:
            workbook = load_workbook(file_path, data_only=True, read_only=True)
        except Exception as exc:
            raise ParsingError(f"Failed to read Excel file {file_path}: {exc}.") from exc

        try:
            items: list[Item] = []
            for worksheet in workbook.worksheets:
                rows = list(worksheet.iter_rows(values_only=True))
                if not rows:
                    continue
                items.extend(_parse_table_rows(rows, source=str(file_path)))
            if not items:
                raise ParsingError(
                    f"No item rows with code, quantity, and price were found in {file_path}."
                )
            return items
        finally:
            workbook.close()

    def parse_pdf(self, file_path: str | Path) -> list[Item]:
        """Parse invoice or packing-list rows from PDF text."""
        try:
            from pypdf import PdfReader
        except ImportError as exc:  # pragma: no cover - dependency normally installed
            raise ParsingError("PDF parsing requires the 'pypdf' package.") from exc

        try:
            reader = PdfReader(str(file_path))
            page_text = [page.extract_text() or "" for page in reader.pages]
        except Exception as exc:
            raise ParsingError(f"Failed to read PDF file {file_path}: {exc}.") from exc

        text = "\n".join(page_text)
        if not text.strip():
            raise ParsingError(f"No extractable text was found in PDF file {file_path}.")

        rows = [line.split() for line in text.splitlines() if line.strip()]
        items = _parse_table_rows(rows, source=str(file_path))
        if not items:
            items = _parse_pdf_lines_without_headers(text.splitlines(), source=str(file_path))
        if not items:
            raise ParsingError(
                f"No item rows with code, quantity, and price were found in {file_path}."
            )
        return items


def parse_document(file_path: str | Path) -> list[Item]:
    """Convenience function to parse a single document."""
    return DocumentParser().parse(file_path)


def _parse_table_rows(rows: Iterable[Iterable[Any]], source: str) -> list[Item]:
    materialized = [list(row) for row in rows]
    for header_index, row in enumerate(materialized):
        mapping = _detect_columns(row)
        if mapping is None:
            continue
        return _items_from_rows(
            materialized[header_index + 1 :], mapping=mapping, source=source
        )
    return []


def _detect_columns(row: list[Any]) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        normalized = _normalize_header(value)
        if not normalized:
            continue
        if normalized in _CODE_HEADERS and "code" not in mapping:
            mapping["code"] = index
        elif normalized in _QUANTITY_HEADERS and "quantity" not in mapping:
            mapping["quantity"] = index
        elif normalized in _PRICE_HEADERS and "price" not in mapping:
            mapping["price"] = index
        elif normalized in _DESCRIPTION_HEADERS and "description" not in mapping:
            mapping["description"] = index
    if {"code", "quantity", "price"}.issubset(mapping):
        return mapping
    return None


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
        .replace("/", "")
        .replace(".", "")
        .replace(":", "")
    )


def _items_from_rows(
    rows: Iterable[list[Any]], mapping: dict[str, int], source: str
) -> list[Item]:
    items: list[Item] = []
    max_index = max(mapping.values())
    for row_number, row in enumerate(rows, start=1):
        if len(row) <= max_index:
            continue
        code_raw = row[mapping["code"]]
        quantity_raw = row[mapping["quantity"]]
        price_raw = row[mapping["price"]]
        if _is_blank(code_raw) and _is_blank(quantity_raw) and _is_blank(price_raw):
            continue
        if _is_summary_row(code_raw):
            continue
        if _is_blank(code_raw) or _is_blank(quantity_raw) or _is_blank(price_raw):
            continue
        try:
            item = Item(
                code=str(code_raw).strip(),
                quantity=parse_decimal(quantity_raw, field_name=f"quantity row {row_number}"),
                price=parse_decimal(price_raw, field_name=f"price row {row_number}"),
                description=(
                    str(row[mapping["description"]]).strip()
                    if "description" in mapping
                    and len(row) > mapping["description"]
                    and not _is_blank(row[mapping["description"]])
                    else None
                ),
            )
        except ParsingError as exc:
            raise ParsingError(f"Failed parsing {source}, row {row_number}: {exc}") from exc
        items.append(item)
    return items


def _parse_pdf_lines_without_headers(lines: Iterable[str], source: str) -> list[Item]:
    """Fallback for simple whitespace-aligned PDFs with code qty price columns."""
    items: list[Item] = []
    for row_number, line in enumerate(lines, start=1):
        parts = line.split()
        if len(parts) < 3:
            continue
        # Find first token containing letters/digits as code, then parse the
        # last two tokens as quantity and price. This handles common simple PDF
        # extracts: "ABC-01 10 2.50" or "ABC-01 goods name 10 2.50".
        code = parts[0]
        quantity_raw = parts[-2]
        price_raw = parts[-1]
        if _normalize_header(code) in _CODE_HEADERS or _is_summary_row(code):
            continue
        try:
            quantity = parse_decimal(quantity_raw, field_name=f"quantity row {row_number}")
            price = parse_decimal(price_raw, field_name=f"price row {row_number}")
        except ParsingError:
            continue
        items.append(Item(code=code.strip(), quantity=quantity, price=price))
    if not items:
        raise ParsingError(f"Unable to identify tabular item data in {source}.")
    return items


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _is_summary_row(value: Any) -> bool:
    normalized = _normalize_header(value)
    return normalized in {"total", "subtotal", "合计", "总计", "小计", "итого"}
