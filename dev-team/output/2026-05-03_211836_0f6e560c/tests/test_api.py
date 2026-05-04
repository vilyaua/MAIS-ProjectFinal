"""Tests for the Supply Chain Management REST API."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from src.app import DEFAULT_API_KEY, create_app, init_db


@pytest.fixture()
def client(tmp_path: Path) -> Any:
    app = create_app({"TESTING": True, "DATABASE": str(tmp_path / "test.db")})
    init_db(app)
    return app.test_client()


@pytest.fixture()
def headers() -> dict[str, str]:
    return {"X-API-Key": DEFAULT_API_KEY}


def post_json(
    client: Any, headers: dict[str, str], path: str, payload: dict[str, Any]
) -> dict[str, Any]:
    response = client.post(path, json=payload, headers=headers)
    assert response.status_code == 201, response.get_json()
    return response.get_json()["data"]


def seed_sku_supplier(
    client: Any, headers: dict[str, str]
) -> tuple[dict[str, Any], dict[str, Any]]:
    sku = post_json(
        client,
        headers,
        "/skus",
        {"code": "SKU-1", "name": "Widget", "unit_price": 12.5},
    )
    supplier = post_json(
        client,
        headers,
        "/suppliers",
        {"name": "Acme", "email": "acme@example.com"},
    )
    return sku, supplier


def test_authentication_is_required(client: Any) -> None:
    response = client.get("/skus")
    assert response.status_code == 401
    assert "invalid API key" in response.get_json()["error"]

    response = client.get("/skus", headers={"X-API-Key": "wrong"})
    assert response.status_code == 401


def test_sku_crud(client: Any, headers: dict[str, str]) -> None:
    create_response = client.post(
        "/skus",
        json={"code": "SKU-CRUD", "name": "CRUD Item", "description": "demo"},
        headers=headers,
    )
    assert create_response.status_code == 201
    sku = create_response.get_json()["data"]

    get_response = client.get(f"/skus/{sku['id']}", headers=headers)
    assert get_response.status_code == 200
    assert get_response.get_json()["data"]["code"] == "SKU-CRUD"

    patch_response = client.patch(
        f"/skus/{sku['id']}", json={"name": "Updated Item"}, headers=headers
    )
    assert patch_response.status_code == 200
    assert patch_response.get_json()["data"]["name"] == "Updated Item"

    delete_response = client.delete(f"/skus/{sku['id']}", headers=headers)
    assert delete_response.status_code == 200
    assert delete_response.get_json()["deleted"] is True

    missing_response = client.get(f"/skus/{sku['id']}", headers=headers)
    assert missing_response.status_code == 404


def test_crud_for_all_domain_entities(client: Any, headers: dict[str, str]) -> None:
    sku, supplier = seed_sku_supplier(client, headers)
    contract = post_json(
        client,
        headers,
        "/contracts",
        {
            "sku_id": sku["id"],
            "supplier_id": supplier["id"],
            "start_date": "2024-01-01",
            "end_date": "2024-12-31",
            "price": 10.0,
        },
    )
    container = post_json(
        client,
        headers,
        "/containers",
        {
            "container_number": "CONT-1",
            "sku_id": sku["id"],
            "supplier_id": supplier["id"],
            "quantity": 100,
            "produced_date": "2024-02-01",
        },
    )
    invoice = post_json(
        client,
        headers,
        "/invoices",
        {
            "invoice_number": "INV-1",
            "supplier_id": supplier["id"],
            "container_id": container["id"],
            "amount": 1000.0,
            "invoice_date": "2024-03-01",
        },
    )
    payment = post_json(
        client,
        headers,
        "/payments",
        {
            "invoice_id": invoice["id"],
            "amount": 500.0,
            "payment_date": "2024-03-15",
            "method": "wire",
        },
    )

    for path, item in (
        ("/contracts", contract),
        ("/containers", container),
        ("/invoices", invoice),
        ("/payments", payment),
    ):
        response = client.get(f"{path}/{item['id']}", headers=headers)
        assert response.status_code == 200
        assert response.get_json()["data"]["id"] == item["id"]


def test_container_filters_by_status_and_date(client: Any, headers: dict[str, str]) -> None:
    sku, supplier = seed_sku_supplier(client, headers)
    post_json(
        client,
        headers,
        "/containers",
        {
            "container_number": "CONT-FILTER-1",
            "sku_id": sku["id"],
            "supplier_id": supplier["id"],
            "quantity": 100,
            "produced_date": "2024-01-05",
        },
    )
    second = post_json(
        client,
        headers,
        "/containers",
        {
            "container_number": "CONT-FILTER-2",
            "sku_id": sku["id"],
            "supplier_id": supplier["id"],
            "quantity": 50,
            "status": "shipped",
            "produced_date": "2024-02-10",
        },
    )

    response = client.get("/containers?status=shipped", headers=headers)
    assert response.status_code == 200
    data = response.get_json()["data"]
    assert [row["id"] for row in data] == [second["id"]]

    response = client.get(
        "/containers?date_field=produced_date&date_from=2024-02-01&date_to=2024-02-28",
        headers=headers,
    )
    assert response.status_code == 200
    data = response.get_json()["data"]
    assert [row["container_number"] for row in data] == ["CONT-FILTER-2"]


def test_container_status_transitions_are_sequential(client: Any, headers: dict[str, str]) -> None:
    sku, supplier = seed_sku_supplier(client, headers)
    container = post_json(
        client,
        headers,
        "/containers",
        {
            "container_number": "CONT-TRANSITION",
            "sku_id": sku["id"],
            "supplier_id": supplier["id"],
            "quantity": 10,
        },
    )

    invalid = client.patch(
        f"/containers/{container['id']}", json={"status": "in transit"}, headers=headers
    )
    assert invalid.status_code == 400
    assert "Invalid container status transition" in invalid.get_json()["error"]

    valid = client.patch(
        f"/containers/{container['id']}", json={"status": "shipped"}, headers=headers
    )
    assert valid.status_code == 200
    assert valid.get_json()["data"]["status"] == "shipped"

    backwards = client.patch(
        f"/containers/{container['id']}", json={"status": "produced"}, headers=headers
    )
    assert backwards.status_code == 400


def test_report_export_returns_excel_workbook(client: Any, headers: dict[str, str]) -> None:
    seed_sku_supplier(client, headers)
    response = client.get("/reports/export", headers=headers)
    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(response.data))
    assert set(workbook.sheetnames) == {
        "skus",
        "suppliers",
        "contracts",
        "containers",
        "invoices",
        "payments",
    }
    assert workbook["skus"].max_row == 2


def test_invalid_input_returns_validation_errors(client: Any, headers: dict[str, str]) -> None:
    response = client.post("/skus", json={"code": ""}, headers=headers)
    assert response.status_code == 400
    body = response.get_json()
    assert body["error"] == "Validation failed"
    assert body["details"]["name"] == "This field is required"

    sku, supplier = seed_sku_supplier(client, headers)
    response = client.post(
        "/containers",
        json={
            "container_number": "BAD-CONT",
            "sku_id": sku["id"],
            "supplier_id": supplier["id"],
            "quantity": 0,
            "produced_date": "not-a-date",
        },
        headers=headers,
    )
    assert response.status_code == 400
    details = response.get_json()["details"]
    assert details["quantity"] == "Must be greater than zero"
    assert "YYYY-MM-DD" in details["produced_date"]
